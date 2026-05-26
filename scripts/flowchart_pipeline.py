#!/usr/bin/env python3
"""
Flowchart Processor — Full Pipeline (v2.2)
============================================
Encapsulates all rules from SKILL.md into a reusable module.
Used by the REST API to process flowchart images.

Supports:
- Blue card detection via connected-component analysis
- Auto/Manual icon detection (center-density method)
- External role labels (above-left of each card)
- Lane-side labels (left column, grouped by y-range)
- System text detection (below card + lane fallback)
- Output: openpyxl xlsx with formatted columns
"""

import os
import io
import numpy as np
from PIL import Image, ImageFilter, ImageEnhance
import pytesseract
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ═══════════════════════════════════════════════════════════════════
# Constants
# ═══════════════════════════════════════════════════════════════════

MIN_CARD_WIDTH = 50
MIN_CARD_HEIGHT = 25
MAX_CARD_WIDTH = 250  # for standard step cards (lane labels may be wider)
MAX_CARD_HEIGHT = 100
ICON_CROP_HEIGHT_RATIO = 0.55
ICON_CROP_WIDTH_RATIO = 0.35
CENTER_DENSITY_THRESHOLD = 0.06  # > 6% center white = Manual icon

# External role label OCR region (relative to card top-left)
EXT_LABEL_LEFT = -90
EXT_LABEL_RIGHT = 15
EXT_LABEL_TOP = -28
EXT_LABEL_BOTTOM = 3

# Lane label scan region
LANE_LABEL_X_LEFT = 0
LANE_LABEL_X_RIGHT = 160

# System text below card region
SYS_BELOW_OFFSET = 2
SYS_BELOW_HEIGHT = 35
DARK_PIXEL_THRESHOLD = 18  # min dark pixels to consider as text
SYS_DARK_SCAN_ROW_THRESHOLD = 12

# Icon area size
ICON_SIZE = 38

# Output column widths
COL_WIDTH_STEP = 40
COL_WIDTH_SYS = 16
COL_WIDTH_ROLE = 16
COL_WIDTH_AUTO = 18
COL_WIDTH_ACTIVITY = 42

# ═══════════════════════════════════════════════════════════════════
# Blue detection
# ═══════════════════════════════════════════════════════════════════

def is_blue_pixel(r, g, b):
    """Check if pixel is blue. Uses relaxed threshold for lighter blue cards."""
    ri, gi, bi = int(r), int(g), int(b)
    return bi > ri + 30 and bi > gi + 30 and bi > 100


def is_dark_pixel(r, g, b):
    """Check if pixel is dark (for system text below card)."""
    return max(r, g, b) < 100


def is_white_pixel(r, g, b):
    """Check if pixel is white (for icon detection)."""
    return r > 200 and g > 200 and b > 200


# ═══════════════════════════════════════════════════════════════════
# Connected-component card detection
# ═══════════════════════════════════════════════════════════════════

def detect_cards(img):
    """
    Detect all blue step cards using connected-component analysis.
    Returns list of dicts: {x1, y1, x2, y2, width, height}
    """
    pixels = np.array(img)
    h, w = pixels.shape[:2]

    # Create blue mask
    blue_mask = np.zeros((h, w), dtype=np.uint8)
    for y in range(0, h, 2):  # step by 2 for speed
        for x in range(0, w, 2):
            r, g, b = pixels[y, x][:3]
            if is_blue_pixel(r, g, b):
                blue_mask[y, x] = 255

    # BFS flood-fill to find connected components
    visited = np.zeros((h, w), dtype=bool)
    cards = []

    for y in range(0, h, 2):
        for x in range(0, w, 2):
            if blue_mask[y, x] == 255 and not visited[y, x]:
                # BFS
                stack = [(x, y)]
                xs, ys = [x], [y]
                visited[y, x] = True
                while stack:
                    cx, cy = stack.pop()
                    for dx, dy in [(2, 0), (-2, 0), (0, 2), (0, -2),
                                   (2, 2), (-2, 2), (2, -2), (-2, -2)]:
                        nx, ny = cx + dx, cy + dy
                        if 0 <= nx < w and 0 <= ny < h:
                            if blue_mask[ny, nx] == 255 and not visited[ny, nx]:
                                visited[ny, nx] = True
                                stack.append((nx, ny))
                                xs.append(nx)
                                ys.append(ny)

                min_x, max_x = min(xs), max(xs)
                min_y, max_y = min(ys), max(ys)
                cw, ch = max_x - min_x, max_y - min_y

                # Filter: only step cards (exclude lane labels, annotations)
                if cw >= MIN_CARD_WIDTH and ch >= MIN_CARD_HEIGHT and cw <= MAX_CARD_WIDTH and ch <= MAX_CARD_HEIGHT:
                    cards.append({
                        "x1": int(min_x), "y1": int(min_y),
                        "x2": int(max_x), "y2": int(max_y),
                        "width": int(cw), "height": int(ch)
                    })

    # Sort by X then Y
    cards.sort(key=lambda c: (c["x1"], c["y1"]))
    return cards


# ═══════════════════════════════════════════════════════════════════
# OCR helpers
# ═══════════════════════════════════════════════════════════════════

def ocr_region(img, x1, y1, x2, y2, psm=6, zoom=4):
    """OCR a region with zoom and thresholding."""
    w, h = img.size
    x1c = max(0, x1)
    y1c = max(0, y1)
    x2c = min(w, x2)
    y2c = min(h, y2)

    if x1c >= x2c or y1c >= y2c:
        return ""

    crop = img.crop((x1c, y1c, x2c, y2c))

    # Convert to grayscale + threshold
    gray = crop.convert("L")
    enhancer = ImageEnhance.Contrast(gray)
    gray = enhancer.enhance(2.0)
    gray = gray.filter(ImageFilter.SHARPEN)
    thresh = gray.point(lambda p: 0 if p < 160 else 255)

    # Zoom
    zw, zh = thresh.size
    thresh = thresh.resize((zw * zoom, zh * zoom), Image.NEAREST)

    try:
        text = pytesseract.image_to_string(
            thresh, lang="eng",
            config=f"--psm {psm} --oem 3 -c tessedit_char_whitelist='ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789 -/'"
        ).strip()
    except Exception:
        text = ""

    # Clean garbled chars
    text = text.replace("\n", " ").replace("  ", " ").strip()
    # Remove lines that are just noise
    lines = text.split(" ")
    cleaned = [l for l in lines if not all(c in "|-—_" for c in l)]
    return " ".join(cleaned)


def ocr_card_text(img, x1, y1, x2, y2):
    """OCR white text inside a blue card. Returns cleaned title + full text."""
    crop = img.crop((x1 + 2, y1 + 2, x2 - 2, y2 - 2))

    # Keep white text pixels, black background
    pixels = np.array(crop.convert("RGB"))
    h, w = pixels.shape[:2]

    # Create inverted image for OCR
    text_img = np.ones((h, w), dtype=np.uint8) * 255  # white bg
    for y in range(h):
        for x in range(w):
            r, g, b = pixels[y, x]
            if is_white_pixel(r, g, b):
                text_img[y, x] = 0  # black text on white bg

    img_pil = Image.fromarray(text_img)
    zw, zh = w * 4, h * 4
    img_pil = img_pil.resize((zw, zh), Image.NEAREST)

    try:
        for thresh_val in [190, 170, 150]:
            thresh = img_pil.point(lambda p: 0 if p < thresh_val else 255)
            text = pytesseract.image_to_string(
                thresh, lang="eng", config="--psm 6 --oem 3"
            ).strip()
            text = text.replace("\n", " ").replace("  ", " ").strip()
            if len(text) > 3:
                break
    except Exception:
        text = ""

    lines = text.split("  ") if "  " in text else [text]
    title = lines[0].strip() if lines else text

    return title, text


# ═══════════════════════════════════════════════════════════════════
# Icon detection (Auto / Manual / User)
# ═══════════════════════════════════════════════════════════════════

def detect_icon(img, x1, y1):
    """Detect card icon type using center-density method."""
    crop = img.crop((x1 + 2, y1 + 2, x1 + ICON_SIZE + 2, y1 + ICON_SIZE + 2))
    pixels = np.array(crop)
    h, w = pixels.shape[:2]

    white_mask = np.zeros((h, w), dtype=bool)
    for y in range(h):
        for x in range(w):
            r, g, b = pixels[y, x][:3]
            if is_white_pixel(r, g, b):
                white_mask[y, x] = True

    white_count = np.sum(white_mask)
    if white_count < 5:
        return ""  # No icon

    # Center density
    cy, cx = h // 2, w // 2
    center_region = white_mask[cy - 3:cy + 3, cx - 3:cx + 3]
    center_white = np.sum(center_region)
    center_density = center_white / 36.0  # 6x6 = 36 pixels

    if center_density > CENTER_DENSITY_THRESHOLD:
        return "Manual"  # Hand icon: solid center
    else:
        return "Auto"  # Gear icon: hollow center


# ═══════════════════════════════════════════════════════════════════
# System text detection (below card)
# ═══════════════════════════════════════════════════════════════════

def detect_system_below_card(img, x1, x2, y2):
    """
    Detect system text directly below a card.
    Scans for dark pixel rows in region (y2+2 to y2+35).
    """
    pixels = np.array(img)
    h, w = pixels.shape[:2]
    sy = y2 + SYS_BELOW_OFFSET
    ey = min(y2 + SYS_BELOW_HEIGHT, h)

    # Scan for dark pixel concentration
    for y in range(sy, ey):
        dark_count = 0
        for x in range(x1, min(x2, w)):
            r, g, b = pixels[y, x][:3]
            if is_dark_pixel(r, g, b):
                dark_count += 1
        if dark_count > SYS_DARK_SCAN_ROW_THRESHOLD:
            # Found text region - OCR it
            text_region_y1 = y
            text_region_y2 = min(y + 20, ey)
            raw = ocr_region(img, x1, text_region_y1, x2, text_region_y2, psm=7, zoom=4)
            return clean_system_text(raw)

    return ""


def clean_system_text(text):
    """Clean garbled OCR output for system names."""
    if not text:
        return ""

    # Check for separator lines
    if all(c in "—-—" for c in text):
        return ""

    # Remove single char fragments
    parts = text.split()
    cleaned = [p for p in parts if len(p) > 1]
    return " ".join(cleaned)


# ═══════════════════════════════════════════════════════════════════
# Lane labels (left side)
# ═══════════════════════════════════════════════════════════════════

def detect_lane_labels(img):
    """
    Detect lane labels on the left side of the flowchart.
    Scans x=0~160 at 15px y-steps.
    Returns dict: {y_start: (label_text, y_end)}
    """
    pixels = np.array(img)
    h, w = pixels.shape[:2]
    right = min(LANE_LABEL_X_RIGHT, w)

    labels = {}
    current_label = None
    label_start = None

    for y in range(0, h, 15):
        # Check if this row has text in the lane label area
        text = ocr_region(img, LANE_LABEL_X_LEFT, y,
                          right, min(y + 20, h), psm=6, zoom=3)

        if text and len(text) > 2:
            cleaned = text.strip()
            if cleaned != current_label:
                if current_label and label_start is not None:
                    labels[label_start] = (current_label, y)
                current_label = cleaned
                label_start = y
        elif current_label and label_start is not None:
            # End of label block
            labels[label_start] = (current_label, y)
            current_label = None
            label_start = None

    if current_label and label_start is not None:
        labels[label_start] = (current_label, h)

    return labels


# ═══════════════════════════════════════════════════════════════════
# Main pipeline
# ═══════════════════════════════════════════════════════════════════

def process_flowchart(image_path_or_bytes):
    """
    Full flowchart processing pipeline.

    Args:
        image_path_or_bytes: Path to image file OR bytes of image

    Returns:
        io.BytesIO: Excel file as bytes buffer
    """
    # Load image
    if isinstance(image_path_or_bytes, bytes):
        img = Image.open(io.BytesIO(image_path_or_bytes))
    else:
        img = Image.open(image_path_or_bytes)

    if img.mode == "RGBA":
        img = img.convert("RGB")

    w, h = img.size
    print(f"[Pipeline] Image: {w}x{h}")

    # Step 1: Detect cards
    cards = detect_cards(img)
    print(f"[Pipeline] Detected {len(cards)} cards")

    if not cards:
        raise ValueError("No step cards detected in image. Check if image contains blue flowchart cards.")

    # Step 2: Detect lane labels
    lane_labels = detect_lane_labels(img)
    print(f"[Pipeline] Lane labels: {len(lane_labels)}")

    # Step 3: Process each card
    results = []
    for i, card in enumerate(cards):
        x1, y1, x2, y2 = card["x1"], card["y1"], card["x2"], card["y2"]

        # Card text
        title, full_text = ocr_card_text(img, x1, y1, x2, y2)
        l4 = title if title else full_text[:60]

        # Icon detection
        auto_val = detect_icon(img, x1, y1)

        # External role label (above-left of card)
        ext_label = ocr_region(
            img,
            x1 + EXT_LABEL_LEFT, y1 + EXT_LABEL_TOP,
            x1 + EXT_LABEL_RIGHT, y1 + EXT_LABEL_BOTTOM,
            psm=7, zoom=4
        )
        ext_label = ext_label.strip() if ext_label else ""

        # System text below card
        system = detect_system_below_card(img, x1, x2, y2)

        print(f"  Card {i+1}: L4='{l4[:30]}' System='{system}' Label='{ext_label}' Auto='{auto_val}'")

        results.append({
            "x1": x1, "y1": y1, "x2": x2, "y2": y2,
            "L4": l4,
            "L5": full_text if full_text else l4,
            "System": system,
            "Auto": auto_val,
            "ExtLabel": ext_label,
            "Role": "",  # Filled in step 4
        })

    # Step 4: Determine Roles (5-level priority)
    # 4a: Build lane labels into y-ranges
    lane_ranges = []
    lane_y_starts = sorted(lane_labels.keys())
    for i, ys in enumerate(lane_y_starts):
        label_text, ye = lane_labels[ys]
        next_ys = lane_y_starts[i + 1] if i + 1 < len(lane_y_starts) else h
        lane_ranges.append({"y1": ys, "y2": next_ys, "label": label_text})

    # 4b: Assign roles
    # Priority 1: External labels
    for r in results:
        if r["ExtLabel"] and len(r["ExtLabel"]) > 1:
            r["Role"] = r["ExtLabel"]

    # Priority 2: Lane labels
    for r in results:
        if not r["Role"]:
            for lr in lane_ranges:
                if lr["y1"] <= r["y1"] + 10 <= lr["y2"]:
                    r["Role"] = lr["label"]
                    break

    # Priority 3: Same-lane inheritance (inherit from labeled card in same y-range)
    for r in results:
        if not r["Role"]:
            for lr in lane_ranges:
                if lr["y1"] <= r["y1"] + 10 <= lr["y2"]:
                    # Find a labeled card in the same lane
                    for other in results:
                        if other["Role"] and lr["y1"] <= other["y1"] + 10 <= lr["y2"]:
                            r["Role"] = other["Role"]
                            break
                    break

    # Priority 4: TBD
    for r in results:
        if not r["Role"]:
            r["Role"] = "TBD"

    # Fill in System from lane labels when below-card text is empty
    for r in results:
        if not r["System"]:
            for lr in lane_ranges:
                if lr["y1"] <= r["y1"] + 10 <= lr["y2"]:
                    r["System"] = lr["label"]
                    break
        if not r["System"]:
            r["System"] = "Offline"

    print(f"[Pipeline] Processed {len(results)} cards successfully")
    return results


# ═══════════════════════════════════════════════════════════════════
# Excel output
# ═══════════════════════════════════════════════════════════════════

def create_excel(cards_data, output_buf=None):
    """
    Generate formatted Excel file.

    Args:
        cards_data: List of card dicts from process_flowchart()
        output_buf: BytesIO, file path (str), or None. If None, returns BytesIO.
                    If str, saves to that path and returns the path.

    Returns:
        io.BytesIO or str (file path)
    """
    is_path = isinstance(output_buf, str)
    if output_buf is None:
        output_buf = io.BytesIO()

    wb = openpyxl.Workbook()
    ws = wb.active

    # Styles
    hdr_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    auto_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
    manual_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
    cell_font = Font(name="Arial", size=10)
    wrap = Alignment(wrap_text=True, vertical="center", horizontal="left")
    center_wrap = Alignment(wrap_text=True, vertical="center", horizontal="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Title row
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = f"Flowchart Extraction — {len(cards_data)} Steps"
    title_cell.font = Font(name="Arial", bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 45

    # Header row
    headers = ["Step", "System", "Role", "Automated or Manual", "Activity"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center_wrap
        cell.border = thin_border
    ws.row_dimensions[2].height = 25

    # Data rows
    for idx, cd in enumerate(cards_data):
        row = idx + 3
        for col_idx, key in enumerate(["L4", "System", "Role", "Auto", "L5"], 1):
            cell = ws.cell(row=row, column=col_idx, value=cd[key])
            cell.font = cell_font
            cell.alignment = wrap
            cell.border = thin_border

        # Color code Auto column
        auto_cell = ws.cell(row=row, column=4)
        if auto_cell.value == "Auto":
            auto_cell.fill = auto_fill
        elif auto_cell.value == "Manual":
            auto_cell.fill = manual_fill

        ws.row_dimensions[row].height = 45

    # Column widths
    ws.column_dimensions["A"].width = COL_WIDTH_STEP
    ws.column_dimensions["B"].width = COL_WIDTH_SYS
    ws.column_dimensions["C"].width = COL_WIDTH_ROLE
    ws.column_dimensions["D"].width = COL_WIDTH_AUTO
    ws.column_dimensions["E"].width = COL_WIDTH_ACTIVITY

    # Freeze header + auto-filter
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:E{len(cards_data) + 2}"

    wb.save(output_buf)
    if is_path:
        return output_buf
    output_buf.seek(0)
    return output_buf


# ═══════════════════════════════════════════════════════════════════
# CLI support
# ═══════════════════════════════════════════════════════════════════

def main():
    import sys
    if len(sys.argv) < 2:
        print("Usage: python3 flowchart_pipeline.py <image_path> [output_path]")
        sys.exit(1)

    image_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else \
        os.path.splitext(image_path)[0] + "_output.xlsx"

    results = process_flowchart(image_path)
    create_excel(results, output_path)
    print(f"\n✅ Excel saved: {output_path}")


if __name__ == "__main__":
    main()
