#!/usr/bin/env python3
"""
Flowchart → Excel extraction script.

⚠️ DEPRECATED: This script only supports 2×3 grid flowcharts with hardcoded coordinates.
Use `flowchart_pipeline.py` instead for arbitrary swimlane/blue-detection flowcharts.

This file is kept for reference only.

Output columns: Step, System, Role, Automated or Manual, Activity

Usage:
    python3 flowchart_to_excel.py <image_path> [output_path]
"""

import sys
import os
from PIL import Image, ImageFilter, ImageEnhance
import pytesseract
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# ——————— Tesseract path ———————
# For PyInstaller bundle support
if getattr(sys, '_MEIPASS', None):
    _bundle_dir = sys._MEIPASS
else:
    _bundle_dir = os.path.dirname(os.path.abspath(__file__))

TESSERACT_CANDIDATES = [
    os.path.join(_bundle_dir, 'tesseract', 'tesseract.exe'),
    os.path.join(_bundle_dir, 'tesseract-ocr', 'tesseract.exe'),
]
for _c in TESSERACT_CANDIDATES:
    if os.path.isfile(_c):
        _td = os.path.join(os.path.dirname(_c), 'tessdata')
        if os.path.isdir(_td):
            os.environ['TESSDATA_PREFIX'] = _td
        pytesseract.pytesseract.tesseract_cmd = _c
        break

# ——————— Card regions ———————
CARD_REGIONS = [
    ("Card 1 (Top-Left)",     310, 120, 440, 225),
    ("Card 2 (Top-Mid)",      480, 120, 615, 225),
    ("Card 3 (Top-Right)",    660, 120, 800, 225),
    ("Card 4 (Bot-Left)",     310, 240, 440, 340),
    ("Card 5 (Bot-Mid)",      480, 240, 615, 340),
    ("Card 6 (Bot-Right)",    660, 240, 800, 340),
]

LEFT_LANE_REGIONS = [
    ("Left Lane 1", 86, 55, 310, 215),
    ("Left Lane 2", 86, 215, 310, 335),
]

# ——————— Role keywords ———————
ROLE_KEYWORDS = [
    "CSM", "Customer Service Manager",
    "CRM", "Customer Relationship Manager",
    "AM", "Account Manager",
    "PM", "Product Manager", "Project Manager",
    "Sales", "Sales Rep",
    "Support", "Support Agent",
    "Admin", "Administrator",
    "Manager", "Supervisor",
    "Analyst", "Data Analyst",
    "Engineer", "Developer",
    "QA", "Tester",
    "Designer", "UI", "UX",
    "Leader", "Team Lead",
    "Director", "VP", "Head of",
    "Agent", "Operator",
    "Bot", "Automation",
    "System", "System Auto",
]


def load_image(path):
    """Load and preprocess image."""
    img = Image.open(path)
    if img.mode == "RGBA":
        img = img.convert("RGB")
    return img


def preprocess_for_ocr(img, x1, y1, x2, y2):
    """Preprocess a card region for better OCR accuracy."""
    crop = img.crop((x1, y1, x2, y2))
    gray = crop.convert("L")
    enhancer = ImageEnhance.Contrast(gray)
    gray = enhancer.enhance(1.5)
    gray = gray.filter(ImageFilter.SHARPEN)
    thresh = gray.point(lambda p: 0 if p < 200 else 255)
    return thresh


def ocr_text_with_data(img, x1, y1, x2, y2):
    """Get OCR text with position data for a region."""
    processed = preprocess_for_ocr(img, x1, y1, x2, y2)
    data = pytesseract.image_to_data(processed, lang="eng", config="--psm 6",
                                     output_type=pytesseract.Output.DICT)
    texts = []
    for i in range(len(data["text"])):
        t = data["text"][i].strip()
        if t and len(t) > 1 and data["conf"][i] > 5:
            texts.append({
                "text": t,
                "x": data["left"][i] + x1,
                "y": data["top"][i] + y1,
                "conf": data["conf"][i],
            })
    # Deduplicate
    if texts:
        deduped = [texts[0]]
        for t in texts[1:]:
            last = deduped[-1]
            if t["text"] == last["text"] and abs(t["y"] - last["y"]) < 5:
                continue
            deduped.append(t)
        texts = deduped
    return texts


def is_role_explicit(text_items):
    """Check if any card text explicitly contains role keywords."""
    full_text = " ".join(item["text"] for item in text_items)
    for kw in ROLE_KEYWORDS:
        if kw.lower() in full_text.lower():
            return True, kw
    return False, ""


def infer_role_from_context(l4_text, system_text):
    """Infer role from step context patterns."""
    text = (l4_text + " " + system_text).lower()
    patterns = [
        (["complaint", "feedback", "inquiry", "request"], "Customer Service Rep"),
        (["escalat", "issue", "ticket"], "CSM / Escalation Manager"),
        (["call", "phone", "contact", "reach out"], "Customer Service Rep"),
        (["approve", "approval"], "Approver / Manager"),
        (["notif", "alert", "remind"], "System Notification"),
        (["check", "verify", "validate", "confirm"], "Validator / QC"),
        (["update", "modify", "change", "edit"], "Data Entry Operator"),
        (["process", "handle", "route", "assign"], "Processor / Dispatcher"),
        (["review", "audit"], "Reviewer"),
        (["decide", "decision"], "Decision Maker"),
        (["reject", "decline"], "Approver"),
        (["auto", "automated", "trigger", "webhook", "api"], "System (Auto)"),
        (["report", "summary", "dashboard"], "Reporting / Analyst"),
    ]
    for keywords, role in patterns:
        if any(kw in text for kw in keywords):
            return role + " (Suggest to check)"
    return ""


def extract_card_info(img, x1, y1, x2, y2, card_name=""):
    """Extract L4, System, Role, L5 from a single card region."""
    texts = ocr_text_with_data(img, x1, y1, x2, y2)
    if not texts:
        return {"L4": "", "System": "offline", "Role": "", "L5": ""}

    # Group into lines by y proximity
    sorted_texts = sorted(texts, key=lambda t: t["y"])
    lines, cur = [], [sorted_texts[0]]
    for t in sorted_texts[1:]:
        if abs(t["y"] - cur[-1]["y"]) < 10:
            cur.append(t)
        else:
            lines.append(cur)
            cur = [t]
    lines.append(cur)

    line_texts = []
    for line in lines:
        sorted_in_line = sorted(line, key=lambda t: t["x"])
        lt = " ".join(t["text"] for t in sorted_in_line).strip()
        if lt:
            line_texts.append(lt)

    flat_text = " ".join(line_texts)
    l4 = line_texts[0] if line_texts else flat_text[:80]

    # System from bottom ~25%
    card_h = y2 - y1
    bottom_y = y2 - card_h * 0.25
    bottom_texts = [t["text"] for t in texts if t["y"] > bottom_y]
    seen = set()
    system_parts = []
    for bt in bottom_texts:
        if bt.lower() not in seen:
            seen.add(bt.lower())
            system_parts.append(bt)
    system = " ".join(system_parts) if system_parts else "offline"

    # Role
    explicit, role_val = is_role_explicit(texts)
    role = role_val if explicit else infer_role_from_context(l4, system)

    return {"L4": l4, "System": system, "Role": role, "L5": flat_text}


def extract_flowchart(image_path):
    """Extract all card data from a flowchart image."""
    img = load_image(image_path)
    cards = []
    for name, x1, y1, x2, y2 in CARD_REGIONS:
        info = extract_card_info(img, x1, y1, x2, y2, card_name=name)
        cards.append(info)
    return cards


def create_excel(cards_data, output_path):
    """Generate Excel file from extracted card data."""
    wb = openpyxl.Workbook()
    ws = wb.active

    hdr_font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
    cell_font = Font(name="Arial", size=10)
    wrap = Alignment(wrap_text=True, vertical="center", horizontal="left")
    center_wrap = Alignment(wrap_text=True, vertical="center", horizontal="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws.merge_cells("A1:D1")
    ws["A1"].value = "Flowchart - 流程结构化数据"
    ws["A1"].font = Font(name="Arial", bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = ["L4", "System", "Role", "L5"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center_wrap
        cell.border = thin_border

    for idx, cd in enumerate(cards_data):
        row = idx + 3
        for col_idx, key in enumerate(["L4", "System", "Role", "L5"], 1):
            cell = ws.cell(row=row, column=col_idx, value=cd[key])
            cell.font = cell_font
            cell.alignment = wrap
            cell.border = thin_border
            if row % 2 == 1:
                cell.fill = alt_fill
        ws.row_dimensions[row].height = 60

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 60

    wb.save(output_path)
    return output_path


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 flowchart_to_excel.py <image_path> [output_path]")
        sys.exit(1)

    image_path = sys.argv[1]
    if not os.path.exists(image_path):
        print(f"Error: File not found: {image_path}")
        sys.exit(1)

    output_path = sys.argv[2] if len(sys.argv) > 2 else \
        os.path.splitext(image_path)[0] + ".xlsx"

    print(f"Loading image: {image_path}")
    cards = extract_flowchart(image_path)
    for name, _, _, _, _ in CARD_REGIONS:
        # Just show card index
        pass
    for i, c in enumerate(cards):
        print(f"  Card {i+1}: L4='{c['L4'][:40]}...' System='{c['System']}' Role='{c['Role']}'")

    result = create_excel(cards, output_path)
    print(f"\n✅ Excel saved: {result}")


if __name__ == "__main__":
    main()
